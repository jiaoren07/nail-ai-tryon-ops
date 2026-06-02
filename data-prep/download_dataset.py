"""
下载赛题脱敏数据集图片到 d:\美团AI HACKATHON\dataset\

目录结构：
  dataset/
    hands/       01.png ... 13.png      (13 张配对手图)
    styles/      01_orig.{png|jpg} ... 25_orig.{png|jpg}
                 01_enh.png       ... 25_enh.png
"""
import os
import sys
from pathlib import Path
from urllib.parse import urlparse
import openpyxl
import httpx

XLSX = r"d:\美团AI HACKATHON\命题三美甲评测数据（对外版） (1).xlsx"
OUT = Path(r"d:\美团AI HACKATHON\dataset")

(OUT / "hands").mkdir(parents=True, exist_ok=True)
(OUT / "styles").mkdir(parents=True, exist_ok=True)


def ext(url: str) -> str:
    return urlparse(url).path.split(".")[-1].lower()


def fetch(client: httpx.Client, url: str, target: Path) -> str:
    if target.exists() and target.stat().st_size > 0:
        return "skip"
    r = client.get(url, timeout=30, follow_redirects=True)
    r.raise_for_status()
    target.write_bytes(r.content)
    return f"ok ({len(r.content)//1024} KB)"


def main():
    wb = openpyxl.load_workbook(XLSX, data_only=True)

    jobs = []  # (url, target_path, label)

    # Sheet 1: 手图 (rows 2..14 are the 13 paired entries)
    ws1 = wb["手图"]
    for i, r in enumerate(range(2, 15), start=1):
        hand_url = ws1.cell(r, 1).value
        if hand_url:
            jobs.append((hand_url, OUT / "hands" / f"{i:02d}.{ext(hand_url)}", f"hand {i:02d}"))

    # Sheet 2: 款式图 (rows 2..26 are 25 entries, idx 1..25)
    ws2 = wb["款式图"]
    for r in range(2, 27):
        idx = ws2.cell(r, 1).value
        orig = ws2.cell(r, 2).value
        enh = ws2.cell(r, 3).value
        if idx and orig:
            jobs.append((orig, OUT / "styles" / f"{idx:02d}_orig.{ext(orig)}", f"style {idx:02d} orig"))
        if idx and enh:
            jobs.append((enh, OUT / "styles" / f"{idx:02d}_enh.{ext(enh)}", f"style {idx:02d} enh"))

    print(f"Total {len(jobs)} files to download.\n")

    ok, skip, fail = 0, 0, 0
    with httpx.Client(headers={"User-Agent": "Mozilla/5.0"}) as client:
        for url, path, label in jobs:
            try:
                status = fetch(client, url, path)
                if status == "skip":
                    skip += 1
                    print(f"  [skip] {label}")
                else:
                    ok += 1
                    print(f"  [ ok ] {label} -> {path.name} {status}")
            except Exception as e:
                fail += 1
                print(f"  [FAIL] {label}: {e}")

    print(f"\nDone. ok={ok}  skip={skip}  fail={fail}  total={len(jobs)}")
    return 0 if fail == 0 else 1


if __name__ == "__main__":
    sys.exit(main())
